import openpyxl
from datetime import datetime

budget_wb_name = "LAD Budget - Export"
lnum_wb_name = "LNumDB"
translation_wb_name = "TranslationDB"

budget = openpyxl.load_workbook("{}.xlsx".format(budget_wb_name))['LAD Budget - Export']
lnums = openpyxl.load_workbook("{}.xlsx".format(lnum_wb_name))['Sheet1']
translation = openpyxl.load_workbook("{}.xlsx".format(translation_wb_name), data_only=True)['Product Translation']

wb = openpyxl.Workbook()
sh = wb.active
sh.title = "LAD Budget"


def transfer_lnums():
    l_dict = {}
    for i in range(2, lnums.max_row + 1):
        l_dict[lnums['A{}'.format(i)].value] = [lnums['B{}'.format(i)].value, lnums['C{}'.format(i)].value,
                                                lnums['D{}'.format(i)].value]

    sh['A1'].value = "Store ID"
    sh['B1'].value = "Marketing Manager"
    sh['M1'].value = "Group Leader"
    orphan_keys = []
    for i in range(2, budget.max_row + 1):
        key = budget['A{}'.format(i)].value
        if key in l_dict:
            sh['A{}'.format(i)].value = l_dict[key][0]
            sh['B{}'.format(i)].value = l_dict[key][1]
            sh['M{}'.format(i)].value = l_dict[key][2]
        else:
            sh['A{}'.format(i)].value = "Missing 233"
            if key not in orphan_keys:
                orphan_keys.append(key)

    if orphan_keys:
        print(orphan_keys)


translate_dict = {}


def create_translate_dict():
    for i in range(2, translation.max_row+1):
        try:
            translate_dict[translation['O{}'.format(i)].value.lower()] = translation['P{}'.format(i)].value.split(";")
        except:
            pass


def translate():
    sh['C1'].value = "Business_Area"
    sh['D1'].value = "Functional_Area"
    sh['E1'].value = "Channel"
    sh['F1'].value = "Type"
    sh['G1'].value = "Solution"
    sh['H1'].value = "Details"
    sh['I1'].value = "Vendor"
    sh['J1'].value = "Solution_Name"
    sh['K1'].value = "Charge_Amount"
    sh['L1'].value = "Store_Paid"
    orphan_keys = []
    for i in range(2, budget.max_row + 1):
        key_list = []
        key_list.extend([
            budget['C{}'.format(i)].value.lower(),
            budget['D{}'.format(i)].value.lower(),
            budget['E{}'.format(i)].value.lower(),
            budget['F{}'.format(i)].value.lower(),
            budget['G{}'.format(i)].value.lower()
        ])
        key = ";".join(key_list)

        if key in translate_dict:
            sh['C{}'.format(i)].value = translate_dict[key][0]
            sh['D{}'.format(i)].value = translate_dict[key][1]
            sh['E{}'.format(i)].value = translate_dict[key][2]
            sh['F{}'.format(i)].value = translate_dict[key][3]
            sh['G{}'.format(i)].value = translate_dict[key][4]
            sh['H{}'.format(i)].value = translate_dict[key][5]
            sh['I{}'.format(i)].value = translate_dict[key][6]
            sh['J{}'.format(i)].value = translate_dict[key][7]

        else:
            sh['C{}'.format(i)].value = key
            orphan_keys.append(key)

        # Charge Amount
        if budget['E{}'.format(i)].value == "Co-op" and int(budget['S{}'.format(i)].value) > 0:
            sh['K{}'.format(i)].value = budget['S{}'.format(i)].value * -1
        else:
            sh['K{}'.format(i)].value = budget['S{}'.format(i)].value

        # Store/LAD Paid
        sh['L{}'.format(i)].value = budget['AI{}'.format(i)].value


    if orphan_keys:
        for key in orphan_keys:
            print(key)

    time = datetime.now()
    now = time.strftime("%m.%d.%Y")

    wb.save("LAD Budget - Translation - {}.xlsx".format(now))


if __name__ == "__main__":
    transfer_lnums()
    create_translate_dict()
    translate()
