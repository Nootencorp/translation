import openpyxl
import re
from datetime import datetime
from difflib import get_close_matches

translation_wb_name = "TranslationDB"

wb = openpyxl.load_workbook(filename="{}.xlsx".format(translation_wb_name))
translation = wb['Product Translation']

read_wb = openpyxl.load_workbook(filename="{}.xlsx".format(translation_wb_name), data_only=True)
read_translation = read_wb['Product Translation']

time = datetime.now()
now = time.strftime("%m.%d.%Y")

brands = ['Ford', 'Mercedes', 'BMW', 'CJD', 'Acura', 'Chevy', 'CHR', 'DOD', 'Honda', 'Hyundai', 'Kia', 'Lincoln',
          'Mazda', 'Mini', 'Nissan', 'Subaru', 'Toyota', 'VW'
          ]

print_vendors = ["Missions Media", "Des Moines Register", "Observer Dispatch", "Santa Barbara News-Press",
                 "Our Community", "MAG", "Trade Express", "Autobody News", "The Korea Times", "The Korea Daily",
                 "Times-Herald Record", "Keizer Times", "Christian Creative News", "Allied Business Service",
                 "Herald Standard", "Hazleton Standard Speaker", "Greater Utica Magazine", "N2 Publishing",
                 "NJ Automotive", "Photo Ad Express", "Hispanic Life Magazine", "Billings Area News Group",
                 "Auto Chaser", "Times Standard", "AutoPlus", "NJ Jewish News", "Auto Shopper", "JBER", "Zundfolge",
                 "The Link Newspaper", "Spokane CDA Living Magazine", "Multnomah Athletic Club", "News Review",
                 "Drive Magazine", "Columbia Basin Herald", "Carmel Pine Cone", "Deals and Wheels", "Tusk",
                 "Honlulu Star Advertiser", "Star Advertiser", "Oregon Jewish Life", "Village Green Publication",
                 "Independent Record", "The News Review", "Our Neck of the Woods", "American Classifieds",
                 "Albany Times Union", "Treasure State Lifestyle Magazine", "Ad Sack", "Hearst Communications",
                 "Simply Family Magazine", "Joyful Living Magazine", "Luxury Lifestyle Report", "Rome Sentinel",
                 "Bennington Cinema", "St. Anthony's Church", "Jersey Sporting News", "Trenton Times",
                 "Jonathon Club Magazine", "Sierra Star", "Casper Star Tribune", "GP Chamber Media",
                 "Challenger Community News", "Bergen Record", "Korea Times", "World Journal", "Lahaina News",
                 "Sheridan Press", "Larchmont Chronicle", "Southern Oregon Magazine", "Lifestyle Magazine",
                 "Star Tribune", "Chismes Magazine", "Beaver Life Magazine", "American Legion", "The Tribune",
                 "Marcoa Fairchild Airforce Base", "Broadway in Boise", "Min Nickel", "Hunting and Fishing News",
                 "Chugach Living", "Auto Show Newspaper", "Clovis Hall of Fame", "Family Values Magazine",
                 "Ashland Directory", "Show and Shine", "Neighbors of Northwest Magazine", "Fast Digital Signs",
                 "Concours d'Elegance", "Free Surf Magazine", "Autos and Trucks", "Golf Korea", "BasinLife.com",
                 "The Fil-Am Courier", "Beaver County Times", "Great Falls Tribune", "Orange County Marketplace",
                 "The Messenger", "Dry Cleaners Journal", "The Spectrum", "East Oregonian", "Thrifty Nickel",
                 "Edge Magazine", "Auto Show Magazine", "Parent Magazine", "Buffalo Rocket", "Parts & People",
                 "Truck Connections", "Parts and People", "In Good Health Newspaper", "Excellence Magazine",
                 "Caller Times", "Press Enterprise", "Rockwell Realty Book and Guide", "Basin Life Magazine",
                 "Carmel Magazine", "Deals & Wheels", "Fender Bender", "Greater Utica Mag", "Hunting & Fishing News",
                 "Marcoa Fairchild AFB Guide", "N2 Creekside Publication", "New Jersey Automotive", "SB News Press",
                 "Ratchet+Wrench "
                 ]

outdoor_vendors = {'Park Outdoor': 'Park Outdoor', 'Lamar': 'Lamar Advertising Company',
                   'One St[eo]{1}p': 'Pattison Outdoor Advertising', 'Outdoor Advertising': 'Outdoor Advertising',
                   'Stott': 'Stott Billboard', 'Meadow Digital Board': 'Meadow Digital Board', 'Fiberdyne': 'Fiberdyne',
                   'Screen\s*vision': 'Screenvision', 'Digital Billboard \w+\s\d/\d': 'Digital Dealership System',
                   'Fast Digital Signs': 'Fast Digital Signs', 'Transformation Media': 'Transformation Media',
                   'NCM': 'NCM', 'General Advertising Agency': 'General Advertising Agency',
                   'Bennington Cinema': 'Bennington Cinema', 'Redwood Highway': 'Redwood Highway',
                   'Outfront Media': 'Outfront Media', 'The Medford Center': 'The Medford Center', 'DiMar': 'DiMar',
                   'Way Organized': 'Way Organized', 'Fall Expo Sale': 'Fall Expo Sale',
                   'Singer Billboard': 'Singer Billboard', 'Pocatello Industrial Park': 'Pocatello Industrial Park',
                   'Jordan Creek Mall': 'Jordan Creek Mall', 'Abel': 'Abel Outdoor Advertising',
                   'FML': 'FML Graphic', 'Fan Bash': 'Fan Bash', 'NCM Media': 'NCM Media',
                   'Cooper Vinyl Collision': 'Cooper Vinyl Collision', 'Medford Mustangs': 'Medford Mustangs',
                   }

production_vendors = {'Constellation': 'Constellation', 'TAAA': 'TAAA', 'Delta': 'Delta Group', 'LAD': 'LAD',
                      'MJC Marketing': 'MJC Marketing', 'Giovatto Advertising': 'Giovatto Advertising',
                      'Riverside': 'Riverside Partners', 'Stream': 'Stream Companies', 'Sue Hartford': 'Sue Hartford',
                      'Gonzalez': 'Gonzalez', 'Mitchell Palmer': 'Mitchell Palmer', 'Sheeraz': 'Sheeraz',
                      'Phenomenon': 'Phenomenon', 'Steve Czajkowski': 'Steve Czajkowski', 'SpotX': 'SpotX',
                      'The Ad Dep': 'The Ad Department', 'Foster Martin': 'Foster Martin', 'MAG': 'MAG',
                      'Walsh(\s)*Sheppard': 'Walsh Sheppard', 'Digital(\s)*Air(\s)*Strike': 'Digital Air Strike',
                      'DealerFire': 'DealerFire', 'Dealer(s)*(\s)*United': 'Dealers United', 'LTI': 'LTI Media',
                      'Dealer(\s)*Inspire': 'Dealer Inspire', 'AutoSweet': 'AutoSweet', 'Force': 'Force Marketing',
                      '72(\s)*Advertising': '72 Advertising', 'DDC': 'DDC', '1-Stop': '1-Stop Media',
                      'Billy Merritt': 'Billy Merritt', 'Pure(\s)*Cars': 'PureCars', 'Graphicka': 'Graphicka',
                      'Octane(\s)*Marketing': 'Octane Marketing'
                      }

all_vendors = {'1-Stop(\\s)*Media': '1-Stop Media', '3(\\s)*Birds(\\s)*Marketing': '3 Birds Marketing',
               '4(\\s)*Imprint': '4 Imprint', '72(\\s)*Advertising': '72 Advertising',
               '800(\\s)*Ignition': '800 Ignition', 'AAA': 'AAA', 'AAG': 'AAG',
               'Abel(\\s)*Outdoor(\\s)*Advertising': 'Abel Outdoor Advertising',
               'Absolute(\\s)*Results': 'Absolute Results',
               'Accelerated(\\s)*Dealer(\\s)*Services(\\s)*': 'Accelerated Dealer Services ',
               'Accessories(\\s)*Digital(\\s)*Solution': 'Accessories Digital Solution', 'ACETech': 'ACETech',
               'Acura': 'Acura', 'Ad(\\s)*Leverage': 'Ad Leverage', 'AdPearance': 'AdPearance', 'AdTaxi': 'AdTaxi',
               'AdvantageTec': 'AdvantageTec', 'Adviserly(\\s)*Inc.': 'Adviserly Inc.', 'AdWolf': 'AdWolf',
               'AES(\\s)*Marketing': 'AES Marketing', 'Affinitiv': 'Affinitiv',
               'Affinity(\\s)*Auto(\\s)*Group': 'Affinity Auto Group', 'Agency(\\s)*Alliance': 'Agency Alliance',
               'AHM': 'AHM', 'AIM': 'AIM', 'Albany(\\s)*Times(\\s)*Union': 'Albany Times Union',
               'Aldis(\\s)*Group': 'Aldis Group', 'Alerus(\\s)*Center': 'Alerus Center', 'Altice': 'Altice',
               'Amplify': 'Amplify', 'AMS(\\s)*Texas': 'AMS Texas', 'Apeiron(\\s)*Marketing': 'Apeiron Marketing',
               'ArmedForceDeals.com': 'ArmedForceDeals.com', 'Aspen(\\s)*Automotive': 'Aspen Automotive',
               'Auto(\\s)*Credit(\\s)*Express': 'Auto Credit Express', 'Auto(\\s)*iPacket': 'Auto iPacket',
               'Auto(\\s)*Lead(\\s)*Pro': 'Auto Lead Pro', 'AutoAlert': 'AutoAlert',
               'AutoBond(\\s)*Marketing(\\s)*Plus': 'AutoBond Marketing Plus', 'AutoFi': 'AutoFi',
               'Autofinder': 'Autofinder', 'AutoLand': 'AutoLand', 'AutoLeadStar': 'AutoLeadStar',
               'Autoloop': 'Autoloop', 'AutoMotion': 'AutoMotion',
               'Automotive(\\s)*Internet(\\s)*Media,(\\s)*Inc.': 'Automotive Internet Media, Inc.',
               'automotiveMastermind': 'automotiveMastermind', 'AutomotiveOnly.com': 'AutomotiveOnly.com',
               'AutoMoxie': 'AutoMoxie', 'AutoPoint': 'AutoPoint', 'AutoSweet': 'AutoSweet', 'AutoTrader': 'AutoTrader',
               'AutoUpLink(\\s)*Tech': 'AutoUpLink Tech', 'AutoWeb': 'AutoWeb',
               'Baierl(\\s)*Management(\\s)*Company': 'Baierl Management Company',
               'Bankruptcy(\\s)*Resource(\\s)*Group': 'Bankruptcy Resource Group',
               'Bennington(\\s)*Cinema': 'Bennington Cinema',
               'Billy(\\s)*Merritt(\\s)*Marketing': 'Billy Merritt Marketing', 'Bing': 'Bing',
               'BlueSky(\\s)*Marketing(\\s)*': 'BlueSky Marketing ', 'BMW': 'BMW', 'BOC(\\s)*Partners': 'BOC Partners',
               "Boeing(\\s)*Employees'(\\s)*Credit(\\s)*Union": "Boeing Employees' Credit Union",
               'Boomtown/Giovato': 'Boomtown/Giovato', 'Bradshaw': 'Bradshaw', 'BrandLync': 'BrandLync',
               'Bridge(\\s)*Marketing': 'Bridge Marketing',
               'C(\\s)*to(\\s)*C(\\s)*Design(\\s)*&(\\s)*Print': 'C to C Design & Print',
               'C-4(\\s)*Analytics': 'C-4 Analytics', 'Cadillac': 'Cadillac', 'Caliber': 'Caliber',
               'CallBox': 'CallBox', 'CallDrip': 'CallDrip', 'Callrevu(\\s)*LLC': 'Callrevu LLC',
               'Capital(\\s)*One': 'Capital One', 'Car(\\s)*Media(\\s)*Group': 'Car Media Group',
               'Car(\\s)*Promotions': 'Car Promotions', 'CarFax': 'CarFax', 'CarGurus': 'CarGurus',
               'CarNow': 'CarNow', 'Carphoria': 'Carphoria', 'CarPro': 'CarPro', 'CarQuotes.com': 'CarQuotes.com',
               'Cars.com': 'Cars.com', 'CarsDirect': 'CarsDirect', 'CarterMac': 'CarterMac', 'Catalogs': 'Catalogs',
               'CDK': 'CDK', 'CF(\\s)*Search(\\s)*Marketing': 'CF Search Marketing',
               'Charge(\\s)*Forward(\\s)*Group': 'Charge Forward Group', 'Chrysler': 'Chrysler', 'Cinemark': 'Cinemark',
               'Cinemas(\\s)*6': 'Cinemas 6', 'City(\\s)*Lending': 'City Lending',
               'Clarivoy,(\\s)*Inc.': 'Clarivoy, Inc.', 'Clean(\\s)*Auto(\\s)*Promotions': 'Clean Auto Promotions',
               'Client(\\s)*Command': 'Client Command', 'ClipLab': 'ClipLab', 'CloudOne': 'CloudOne',
               'Coastal(\\s)*Media': 'Coastal Media', 'Coastal(\\s)*Media(\\s)*': 'Coastal Media ',
               'Comcast': 'Comcast', 'Commercial(\\s)*Truck(\\s)*Trader': 'Commercial Truck Trader',
               'Conquest': 'Conquest', 'Constellation': 'Constellation',
               'Consumer(\\s)*Portfolio(\\s)*Services': 'Consumer Portfolio Services', 'Conversica': 'Conversica',
               'Cooper(\\s)*Vinyl(\\s)*Collision': 'Cooper Vinyl Collision', 'Costco': 'Costco',
               'CPS(\\s)*Direct(\\s)*Marketing': 'CPS Direct Marketing', 'Craigslist': 'Craigslist',
               'Creative(\\s)*Mills': 'Creative Mills', 'Creative(\\s)*Package': 'Creative Package',
               'CreditPlus': 'CreditPlus', 'CreditYES': 'CreditYES', 'CU(\\s)*Direct': 'CU Direct',
               'Cypress(\\s)*Alliance(\\s)*': 'Cypress Alliance ', 'Data(\\s)*Clover': 'Data Clover',
               'Day(\\s)*Management(\\s)*Company': 'Day Management Company', 'DDC': 'DDC', 'DDS': 'DDS',
               'Dealer(\\s)*eProcess': 'Dealer eProcess', 'Dealer(\\s)*Fusion': 'Dealer Fusion',
               'Dealer(\\s)*Imaging': 'Dealer Imaging', 'Dealer(\\s)*Inspire': 'Dealer Inspire',
               'Dealer(\\s)*Specialties': 'Dealer Specialties', 'Dealer(\\s)*Teamwork': 'Dealer Teamwork',
               'Dealer(\\s)*Wizard': 'Dealer Wizard', 'DealerAssistNow': 'DealerAssistNow', 'DealerCMO': 'DealerCMO',
               'DealerCustomerMarketing': 'DealerCustomerMarketing', 'DealerFire': 'DealerFire',
               'DealerFuel': 'DealerFuel', 'DealerOn': 'DealerOn', 'DealerRater': 'DealerRater',
               'Dealers(\\s)*Direct(\\s)*USA': 'Dealers Direct USA', 'Dealers(\\s)*United': 'Dealers United',
               'DealerScience': 'DealerScience', 'DealersLink': 'DealersLink', 'DealerTrack': 'DealerTrack',
               'DealerVision': 'DealerVision', 'DealerWorld': 'DealerWorld', 'Delta(\\s)*Group': 'Delta Group',
               'Demand(\\s)*Local': 'Demand Local', 'Detroit(\\s)*Trading(\\s)*Services': 'Detroit Trading Services',
               'DexYP': 'DexYP', 'DH5': 'DH5', 'DiamondLot': 'DiamondLot',
               'Digital(\\s)*Air(\\s)*Strike': 'Digital Air Strike',
               'Digital(\\s)*Dealership(\\s)*System': 'Digital Dealership System',
               'Digital(\\s)*to(\\s)*Dealer(\\s)*Direct': 'Digital to Dealer Direct', 'DiMar': 'DiMar',
               "Dinho's(\\s)*Services": "Dinho's Services", 'Dominion': 'Dominion', 'DPS': 'DPS',
               'Drive(\\s)*Motors': 'Drive Motors', 'Dropzone(\\s)*Tech': 'Dropzone Tech', 'DyGen': 'DyGen',
               'Dynavid': 'Dynavid', 'Ebay': 'Ebay', 'Edifice': 'Edifice', 'Edmunds': 'Edmunds', 'Eleads': 'Eleads',
               'Electra-Media(\\s)*Inc.': 'Electra-Media Inc.', 'Elite(\\s)*Lending': 'Elite Lending',
               'Epsilon': 'Epsilon', 'Expert(\\s)*DMS': 'Expert DMS', 'Facebook': 'Facebook', 'Fair': 'Fair',
               'Fall(\\s)*Expo(\\s)*Sale': 'Fall Expo Sale', 'Fast(\\s)*Digital(\\s)*Signs': 'Fast Digital Signs',
               'FCA': 'FCA', 'Fiberdyne': 'Fiberdyne', 'Fixed(\\s)*Ops(\\s)*Digital': 'Fixed Ops Digital',
               'FMG(\\s)*Print(\\s)*Solutions': 'FMG Print Solutions', 'FML(\\s)*Marketing': 'FML Marketing',
               'Force(\\s)*Marketing': 'Force Marketing', 'Ford': 'Ford', 'Ford(\\s)*Direct': 'Ford Direct',
               'Fortibus': 'Fortibus', 'Foster(\\s)*Automotive': 'Foster Automotive',
               'Foster(\\s)*Martin': 'Foster Martin', 'Fox(\\s)*26': 'Fox 26', 'Fox(\\s)*Dealer': 'Fox Dealer',
               'Free(\\s)*Surf(\\s)*Magazine': 'Free Surf Magazine', 'Fresno(\\s)*Bee': 'Fresno Bee',
               'Friendemic': 'Friendemic', 'Fueled(\\s)*Automotive(\\s)*Marketing': 'Fueled Automotive Marketing',
               'fusionZONE(\\s)*Automotive': 'fusionZONE Automotive', 'Galaxy(\\s)*Radio': 'Galaxy Radio',
               'Gary(\\s)*Stock(\\s)*Company': 'Gary Stock Company', 'GateHouse(\\s)*Media': 'GateHouse Media',
               'General(\\s)*Advertising(\\s)*Agency': 'General Advertising Agency',
               'GGP(\\s)*Clackamas(\\s)*Town(\\s)*Center': 'GGP Clackamas Town Center',
               'Giovatto(\\s)*Advertising': 'Giovatto Advertising', 'GM': 'GM', 'Go(\\s)*Daddy': 'Go Daddy',
               'Gold(\\s)*Digger': 'Gold Digger', 'Gonzalez': 'Gonzalez', 'Goodway(\\s)*Group': 'Goodway Group',
               'Google': 'Google', 'Graphicka': 'Graphicka', 'Great(\\s)*Falls(\\s)*Tribune': 'Great Falls Tribune',
               'GrooveCar': 'GrooveCar', 'Guardian': 'Guardian', 'Gubagoo': 'Gubagoo',
               'Hamlin(\\s)*&(\\s)*Associates': 'Hamlin & Associates', 'Hammer': 'Hammer',
               'Hearst(\\s)*Communications': 'Hearst Communications', 'Herald(\\s)*&(\\s)*News': 'Herald & News',
               'Hill(\\s)*Marketing': 'Hill Marketing', 'HomeNet': 'HomeNet', 'Honda': 'Honda',
               'Honda(\\s)*DCMS': 'Honda DCMS', 'Hopkins(\\s)*and(\\s)*Raines': 'Hopkins and Raines', 'HOX': 'HOX',
               'Hunter(\\s)*Communications': 'Hunter Communications', 'Hype(\\s)*Media': 'Hype Media',
               'Hyundai': 'Hyundai', 'IAM(\\s)*Sales': 'IAM Sales',
               'Icon(\\s)*Internet(\\s)*Media': 'Icon Internet Media', 'Idaho(\\s)*Mt(\\s)*Express': 'Idaho Mt Express',
               'Ideal(\\s)*Direct(\\s)*Ad(\\s)*Group': 'Ideal Direct Ad Group',
               'Ignition(\\s)*Auto(\\s)*Marketing': 'Ignition Auto Marketing',
               'iHeartMedia,(\\s)*Inc.': 'iHeartMedia, Inc.', 'ILM': 'ILM', 'iMedia(\\s)*Network': 'iMedia Network',
               'iMR': 'iMR', 'InMarketSolutions': 'InMarketSolutions', 'Insignia': 'Insignia',
               'InterActive': 'InterActive', 'Interactive(\\s)*Financial': 'Interactive Financial',
               'Inventory(\\s)*Command(\\s)*Center': 'Inventory Command Center', 'iPreCheck': 'iPreCheck',
               'J(\\s)*&(\\s)*L(\\s)*Marketing': 'J & L Marketing', 'Jeffrey(\\s)*G(\\s)*Bonnell': 'Jeffrey G Bonnell',
               'JenStar': 'JenStar', 'JMG': 'JMG', 'Jordan(\\s)*Creek(\\s)*Mall': 'Jordan Creek Mall', 'KBB': 'KBB',
               'Kennedy(\\s)*Marketing(\\s)*Group': 'Kennedy Marketing Group', 'Kia': 'Kia', 'KMG': 'KMG',
               'KREM.com': 'KREM.com', 'KSL(\\s)*Cars': 'KSL Cars', 'KXTL': 'KXTL', 'L2T(\\s)*Media': 'L2T Media',
               'LAD': 'LAD', 'Lamar(\\s)*Advertising(\\s)*Company': 'Lamar Advertising Company',
               'Lana(\\s)*Lane(\\s)*Studios': 'Lana Lane Studios', 'Laser(\\s)*Stream(\\s)*Media': 'Laser Stream Media',
               'Lendward': 'Lendward', 'Levich(\\s)*Group': 'Levich Group', 'Lithia': 'Lithia',
               'Live(\\s)*Event(\\s)*Stream(\\s)*Automotive': 'Live Event Stream Automotive',
               'Live(\\s)*LA(\\s)*Media': 'Live LA Media', 'LivePerson': 'LivePerson', 'LotBoys': 'LotBoys',
               'LotLinx': 'LotLinx', 'LotVantage': 'LotVantage', 'LT1(\\s)*Media': 'LT1 Media',
               'Luminous(\\s)*Sound': 'Luminous Sound', 'MAG': 'MAG',
               'Main(\\s)*Event(\\s)*Chicago': 'Main Event Chicago', 'MarkMonitor': 'MarkMonitor',
               'Mascor(\\s)*Media': 'Mascor Media', 'MaxDigital': 'MaxDigital',
               'Maximum(\\s)*Performance(\\s)*Group': 'Maximum Performance Group', 'MCE': 'MCE',
               'Meadow(\\s)*Digital(\\s)*Board': 'Meadow Digital Board', 'MediaAMP': 'MediaAMP',
               'Mega(\\s)*Marketing': 'Mega Marketing', 'Member(\\s)*Services': 'Member Services',
               'Mercedes-Benz': 'Mercedes-Benz', 'Midco(\\s)*College(\\s)*Football': 'Midco College Football',
               'Mitchell(\\s)*Palmer': 'Mitchell Palmer', 'MJC(\\s)*Marketing': 'MJC Marketing', 'Modal': 'Modal',
               'Motivated(\\s)*Marketing': 'Motivated Marketing', 'MotoMiner': 'MotoMiner', 'MPG': 'MPG',
               'MUDD(\\s)*Advertising(\\s)*': 'MUDD Advertising ', 'MXS(\\s)*Solutions': 'MXS Solutions',
               'MyDealerOnline': 'MyDealerOnline', 'Naked(\\s)*Lime': 'Naked Lime', 'NCM': 'NCM',
               'NcompassTrac': 'NcompassTrac', 'NetDriven': 'NetDriven', 'Netsertive': 'Netsertive', 'Nissan': 'Nissan',
               'NJ(\\s)*Advanced(\\s)*Media': 'NJ Advanced Media', 'NPR': 'NPR', 'Nusani(\\s)*Media': 'Nusani Media',
               'OCTAGON': 'OCTAGON', 'Octane(\\s)*Leads': 'Octane Leads', 'Octane(\\s)*Marketing': 'Octane Marketing',
               'Optima(\\s)*Automotive': 'Optima Automotive', 'Orbee': 'Orbee',
               'Outdoor(\\s)*Advertising': 'Outdoor Advertising', 'Outfront(\\s)*Media': 'Outfront Media',
               'Outsell(\\s)*Corp': 'Outsell Corp', 'OverTake(\\s)*Digital': 'OverTake Digital', 'Pacer': 'Pacer',
               'Pandora': 'Pandora', 'Park(\\s)*Outdoor': 'Park Outdoor',
               'Passport(\\s)*Unlimited': 'Passport Unlimited',
               'Pattison(\\s)*Outdoor(\\s)*Advertising': 'Pattison Outdoor Advertising',
               'Paxton(\\s)*Automotive(\\s)*Marketing': 'Paxton Automotive Marketing',
               'Peak(\\s)*Performance': 'Peak Performance', 'Perfect(\\s)*Enterprises': 'Perfect Enterprises',
               'PERQ': 'PERQ', 'Phenomenon': 'Phenomenon', 'PHMG': 'PHMG', 'Photo(\\s)*Ventures': 'Photo Ventures',
               'Picture(\\s)*This(\\s)*Media': 'Picture This Media', 'PIP(\\s)*Printing': 'PIP Printing',
               'Pixel(\\s)*Motion': 'Pixel Motion',
               'Platinum(\\s)*Analytics(\\s)*&(\\s)*Distribution': 'Platinum Analytics & Distribution',
               'Pocatello(\\s)*Industrial(\\s)*Park': 'Pocatello Industrial Park', 'PocketExpert': 'PocketExpert',
               'Podium': 'Podium', 'Portland(\\s)*Timbers': 'Portland Timbers',
               'Portland(\\s)*Trail(\\s)*Blazers': 'Portland Trail Blazers',
               'Premium(\\s)*Digital(\\s)*Video': 'Premium Digital Video',
               'Prestige(\\s)*Auto(\\s)*Marketing': 'Prestige Auto Marketing', 'Prime(\\s)*Response': 'Prime Response',
               'Printmoz.com': 'Printmoz.com', 'Pro(\\s)*Motion': 'Pro Motion',
               'Proclaim(\\s)*Promotions': 'Proclaim Promotions', 'PSM(\\s)*Marketing': 'PSM Marketing',
               'Pure(\\s)*Digital(\\s)*TV': 'Pure Digital TV', 'PureCars': 'PureCars',
               'Pureinfluencer.com': 'Pureinfluencer.com', 'Quarius': 'Quarius',
               'Recall(\\s)*Masters': 'Recall Masters', 'Redline(\\s)*Advantage': 'Redline Advantage',
               'Redwood(\\s)*Highway': 'Redwood Highway', 'Reputation.com': 'Reputation.com',
               'Results(\\s)*HQ': 'Results HQ', 'Rev(\\s)*It(\\s)*Up(\\s)*Events': 'Rev It Up Events',
               'Revolution(\\s)*Parts': 'Revolution Parts',
               'RH(\\s)*Automotive(\\s)*Solutions': 'RH Automotive Solutions', 'Rick(\\s)*Buffkin': 'Rick Buffkin',
               'Riverside(\\s)*New(\\s)*Car(\\s)*Dealer(\\s)*Association': 'Riverside New Car Dealer Association',
               'Riverside(\\s)*Partners': 'Riverside Partners', 'RK(\\s)*Credit': 'RK Credit', 'Roadster': 'Roadster',
               'Rodo': 'Rodo', 'Ross(\\s)*Media': 'Ross Media', 'Roy(\\s)*Robinson': 'Roy Robinson',
               'RPM(\\s)*Marketing': 'RPM Marketing', 'S3(\\s)*Solutions': 'S3 Solutions',
               'Sales(\\s)*360': 'Sales 360', 'Santander': 'Santander', 'Say-So': 'Say-So',
               'Scott(\\s)*McFadden(\\s)*Productions': 'Scott McFadden Productions', 'Screenvision': 'Screenvision',
               'Search(\\s)*Optics': 'Search Optics', 'Sensible(\\s)*Driver': 'Sensible Driver', 'Sheeraz': 'Sheeraz',
               'ShiftDigital': 'ShiftDigital', 'Shop(\\s)*Click(\\s)*Drive': 'Shop Click Drive',
               'Shore(\\s)*Good(\\s)*Eats(\\s)*N(\\s)*Treats': 'Shore Good Eats N Treats', 'Sign(\\s)*Pro': 'Sign Pro',
               'Signs(\\s)*Sealed(\\s)*&(\\s)*Delivered': 'Signs Sealed & Delivered',
               'Singer(\\s)*Billboard': 'Singer Billboard', 'SLS': 'SLS', 'SmartSites': 'SmartSites',
               'SMedia': 'SMedia', 'Snap21': 'Snap21',
               'Solutions(\\s)*Direct(\\s)*Marketing': 'Solutions Direct Marketing',
               'Southbay(\\s)*Promotions': 'Southbay Promotions', 'Spectrio': 'Spectrio', 'Spectrum': 'Spectrum',
               'SpinCar': 'SpinCar', 'Spotify': 'Spotify', 'SpotX': 'SpotX', 'Star(\\s)*Advertiser': 'Star Advertiser',
               'Steve(\\s)*Czajkowski': 'Steve Czajkowski', 'Stott(\\s)*Billboard': 'Stott Billboard',
               'Strategic(\\s)*Marketing': 'Strategic Marketing',
               'Strategic(\\s)*Marketing(\\s)*': 'Strategic Marketing ', 'Stream(\\s)*Companies': 'Stream Companies',
               'Subaru': 'Subaru', 'Sue(\\s)*Hartford': 'Sue Hartford', 'TAAA': 'TAAA', 'Tacito': 'Tacito',
               'TAPinto(\\s)*Local': 'TAPinto Local', 'TBD': 'TBD', 'Team(\\s)*Velocity': 'Team Velocity',
               'Tecobi': 'Tecobi', 'Test(\\s)*Drive(\\s)*Promotions': 'Test Drive Promotions',
               'Text2Drive': 'Text2Drive', 'Textonix': 'Textonix', 'The(\\s)*Ad(\\s)*Department': 'The Ad Department',
               'The(\\s)*Appraisal(\\s)*Lane': 'The Appraisal Lane', 'The(\\s)*Frank(\\s)*Agency': 'The Frank Agency',
               'The(\\s)*Medford(\\s)*Center': 'The Medford Center',
               'Third(\\s)*Degree(\\s)*Graphics': 'Third Degree Graphics',
               'This(\\s)*Guy(\\s)*Productions': 'This Guy Productions', 'TireConnect': 'TireConnect',
               'Top(\\s)*Deck(\\s)*Medioa': 'Top Deck Medioa',
               'Total(\\s)*Customer(\\s)*Connect': 'Total Customer Connect', 'Tournament(\\s)*Time': 'Tournament Time',
               'Toyota': 'Toyota', 'TradeMotion': 'TradeMotion', 'TradePending(\\s)*': 'TradePending ',
               'Traffic(\\s)*Jam(\\s)*Events': 'Traffic Jam Events',
               'Transformation(\\s)*Media': 'Transformation Media', 'Trilogy': 'Trilogy', 'TrueCar': 'TrueCar',
               'TurnKey(\\s)*Marketing': 'TurnKey Marketing', 'United(\\s)*Rentals': 'United Rentals',
               'Unity(\\s)*Works': 'Unity Works',
               'Up(\\s)*In(\\s)*The(\\s)*Air(\\s)*Recall(\\s)*Marketing': 'Up In The Air Recall Marketing',
               'Uperator': 'Uperator', 'UrbanScience/AutoHook': 'UrbanScience/AutoHook',
               'US(\\s)*Free(\\s)*Press': 'US Free Press', 'Valassis': 'Valassis', 'Vast': 'Vast',
               'Vehicles(\\s)*TEST(\\s)*Publishing': 'Vehicles TEST Publishing', 'Vertical(\\s)*Guru': 'Vertical Guru',
               'VinSolutions': 'VinSolutions', 'VIPdrv': 'VIPdrv',
               'VistaDash(\\s)*Software,(\\s)*Inc.': 'VistaDash Software, Inc.', 'VNA': 'VNA',
               'Volkswagen': 'Volkswagen', 'Wallbank(\\s)*LLC': 'Wallbank LLC', 'Walsh(\\s)*Sheppard': 'Walsh Sheppard',
               'Way(\\s)*Organized': 'Way Organized', 'Weebly': 'Weebly', 'Westates': 'Westates',
               'WheelsTV': 'WheelsTV', 'Widewail': 'Widewail', 'Work(\\s)*Truck(\\s)*Solutions': 'Work Truck Solutions',
               'Xcel(\\s)*Media(\\s)*Group': 'Xcel Media Group', 'Xcite(\\s)*Advertising': 'Xcite Advertising',
               'Xtime(\\s)*Inc.': 'Xtime Inc.', 'Yelp!': 'Yelp!', 'YouTube': 'YouTube', 'Zonic': 'Zonic',
               'Zyon(\\s)*Marketing': 'Zyon Marketing', 'Car(\\s)*People(\\s)*Marketing': 'Car People Marketing',
               'TagRail': 'TagRail', 'We(\\s)*Drive(\\s)*Auto': 'WeDriveAuto', 'Service(\\s)*Global': 'Service Global',
               'Global(\\s)*Database(\\s)*Marketing': 'Global Database Marketing'
               }


def set_func_area(dept):
    if dept == "Sales":
        return "Sales Advertising"
    else:
        return "Fixed Ops Advertising"


def set_details(campaign_name):
    details = ''
    if re.search(r'Video', str(campaign_name), re.I):
        details = "Video"

    return details


def set_vendor(campaign_name, dmvendor, vendor_dict):
    vendor = None
    for k, v in vendor_dict.items():
        if re.search(r'{}'.format(k), str(dmvendor), re.I):
            vendor = v
        elif re.search(r'{}'.format(k), str(campaign_name), re.I):
            vendor = v
        if vendor == None:
            vendor = "MAG"

    return vendor


def auto_update(campaign_name, dept, ttype, dmvendor):
    entry = None

    if ttype == "Outdoor":
        if re.search(r'Digital', str(campaign_name), re.I):
            product = "Digital Signage"
        elif re.search(r'Cinema', str(campaign_name), re.I) or re.search(r'Movie', str(campaign_name), re.I):
            product = "Movie Theater Ads"
        else:
            product = "Signage"

        entry = ['Marketing', set_func_area(dept), 'Traditional', 'Mass Marketing', product, '',
                 set_vendor(campaign_name, dmvendor, outdoor_vendors), campaign_name
                 ]

    elif ttype == "Print":
        vendor = None
        for print_vendor in print_vendors:
            if re.search(r'{}'.format(print_vendor), str(campaign_name), re.I):
                vendor = print_vendor
            else:
                pass

            if vendor:
                entry = ['Marketing', set_func_area(dept), 'Traditional', 'Mass Marketing', 'Print Advertising', '',
                         vendor, campaign_name
                         ]
            else:
                entry = ['Marketing', set_func_area(dept), 'Traditional', 'Mass Marketing', 'Print Advertising', '',
                         'MAG', campaign_name
                         ]

    # Facebook
    elif re.search(r'(^|\s)Facebook(\s|$|/)', str(campaign_name), re.I):
        # Product
        if re.search(r'(\s)Marketplace(\s|$|)', str(campaign_name), re.I):
            product_category = "Inventory Listings"
            product = "3rd Party Inventory Listings"
            product_subcategory = "Facebook Marketplace"
        elif re.search(r'(\s)Retargeting(\s|$|)', str(campaign_name), re.I):
            product_category = "Social - Paid"
            product = "Retargeting"
            product_subcategory = "Facebook"
        elif re.search(r'(\s)Messenger(\s|$|)', str(campaign_name), re.I):
            product_category = "Social - Organic"
            product = "Social Messaging"
            product_subcategory = "Facebook"
        elif re.search(r'(\s)Posting(\s|$|)', str(campaign_name), re.I):
            product_category = "Social - Organic"
            product = "Social Posting"
            product_subcategory = "Facebook"
        else:
            product_category = "Social - Paid"
            product = "Social Ads"
            product_subcategory = "Facebook"

        entry = ['Marketing', set_func_area(dept), 'Online', product_category, product, product_subcategory,
                 set_vendor(campaign_name, dmvendor, production_vendors), campaign_name
                 ]

    # Offers
    elif re.search(r'Offers', str(campaign_name), re.I):
        mfg = None
        for brand in brands:
            if re.search(r'{}'.format(brand), str(campaign_name), re.I):
                mfg = brand
            else:
                pass
        if str(ttype) == "Television" and not re.search(r'YouTube', str(campaign_name), re.I):
            if mfg:
                entry = ['Marketing', set_func_area(dept), 'Traditional', 'Mass Marketing', 'Television Advertising',
                         '', 'MAG', '{} Offers'.format(mfg)
                         ]
            else:
                entry = ['Marketing', set_func_area(dept), 'Traditional', 'Mass Marketing', 'Television Advertising',
                         '', 'MAG', 'Offers'
                         ]
        elif str(ttype) == "Radio":
            if mfg:
                entry = ['Marketing', set_func_area(dept), 'Traditional', 'Mass Marketing', 'Radio Advertising', '',
                         'MAG', '{} Offers'.format(mfg)
                         ]
            else:
                entry = ['Marketing', set_func_area(dept), 'Traditional', 'Mass Marketing', 'Radio Advertising', '',
                         'MAG', 'Offers'
                         ]

    # Television
    elif re.search(r'(pre)(\s|-)*(rol+)', str(campaign_name), re.I):
        if re.search(r'SpotX', str(campaign_name), re.I):
            entry = ['Marketing', set_func_area(dept), 'Online', 'Display',
                     'Video Pre-roll', '', 'SpotX', 'SpotX Pre-roll'
                     ]
        elif re.search(r'TAAA', str(campaign_name), re.I):
            entry = ['Marketing', set_func_area(dept), 'Online', 'Display', 'Video Pre-roll', '', 'TAAA',
                     'TAAA Pre-roll'
                     ]
        else:
            entry = ['Marketing', set_func_area(dept), 'Online', 'Display', 'Video Pre-roll', '', 'MAG', campaign_name
                     ]
    elif re.search(r'(^|\s)OTT(\s|$|/|:)', str(campaign_name), re.I):
        if re.search(r'Compulse', str(campaign_name), re.I):
            entry = ['Marketing', set_func_area(dept), 'Online', 'Display', 'OTT', '', 'MAG', ':30 Compulse OTT'
                     ]
        elif re.search(r':15', str(campaign_name), re.I):
            entry = ['Marketing', set_func_area(dept), 'Online', 'Display', 'OTT', '', 'MAG', ':15 OTT'
                     ]
        elif re.search(r':NFL', str(campaign_name), re.I):
            entry = ['Marketing', set_func_area(dept), 'Online', 'Display', 'OTT', '', 'MAG', ':15 OTT'
                     ]
        else:
            entry = ['Marketing', set_func_area(dept), 'Online', 'Display',
                     'OTT', '', 'MAG', ':30 OTT'
                     ]
    elif re.search(r'(^|\s)Spectrum(\s|/|:)', str(campaign_name), re.I):
        if re.search(r'(^|\s)Cable(\s|$|/|:)', str(campaign_name), re.I):
            entry = ['Marketing', set_func_area(dept), 'Traditional', 'Mass Marketing',
                     'Television Advertising', '', 'Spectrum', campaign_name
                     ]
        else:
            entry = ['Marketing', set_func_area(dept), 'Online', 'Display',
                     'OTT', '', 'Spectrum', campaign_name
                     ]
    elif ttype == "Television" and not re.search(r'(pre)(\s|-)*(rol+)', str(campaign_name), re.I) and not \
            re.search(r'(^|\s)OTT(\s|$|/|:)', str(campaign_name), re.I) and not \
            re.search(r'(^|\s)YouTube(\s|$|/|:)', str(campaign_name), re.I) and not \
            re.search(r'(^|\s)Spectrum(\s|$|/|:)', str(campaign_name), re.I) and not \
            re.search(r'(^|\s)Facebook(\s|$|/)', str(campaign_name), re.I):
        entry = ['Marketing', set_func_area(dept), 'Traditional', 'Mass Marketing', 'Television Advertising', '',
                 set_vendor(campaign_name, dmvendor, production_vendors), campaign_name
                 ]
    # Cable Advertising
    elif re.search(r'cable', str(campaign_name), re.I) or re.search(r'cabel', str(campaign_name), re.I) or \
            re.search(r'cbl', str(campaign_name), re.I):
        entry = ['Marketing', set_func_area(dept), 'Traditional', 'Mass Marketing', 'Television Advertising', '',
                 set_vendor(campaign_name, dmvendor, production_vendors), campaign_name
                 ]

    # Misc
    elif ttype == "Misc":
        if re.search(r'Appraisal Lane', str(campaign_name), re.I):
            entry = ['Marketing', 'Sales Advertising', 'Online', 'Website', 'Website Add-ons', 'Trade-in Tools',
                     'The Appraisal Lane', 'Live Trade-Back Vendor'
                     ]
        elif re.search(r'Baierl Mgmt Co Credit', str(campaign_name), re.I):
            entry = ['Administrative', 'Accounting', 'Professional Services', '', 'Fee', '',
                     'Baierl Management Company', 'Baierl Management Company Credit'
                     ]
        elif re.search(r'Recovery', str(campaign_name), re.I):
            entry = ['Administrative', 'Accounting', 'Professional Services', '', 'Adjustment', '', 'Lithia',
                     'Co-op Recovery'
                     ]
        elif re.search(r'CU Xpress', str(campaign_name), re.I):
            entry = ['Sales Ops', 'Finance and Insurance', 'Software', '', 'Lending Locator', '', 'CU Direct',
                     'CU Xpress Lease Subscription'
                     ]
        elif re.search(r'Dealer(s)*Link', str(campaign_name), re.I):
            entry = ['Marketing', 'Sales Advertising', 'Online', 'Paid Leads', 'Special Finance Leads', '',
                     'DealersLink', 'DealersLink'
                     ]
        elif re.search(r'Delta', str(campaign_name), re.I):
            if re.search(r'Fee', str(campaign_name), re.I):
                entry = ['Administrative', 'Accounting', 'Professional Services', '', 'Fee', '', 'Delta Group',
                         'Agency Fee'
                         ]
            elif re.search(r'Rebate', str(campaign_name), re.I):
                entry = ['Administrative', 'Accounting', 'Professional Services', '', 'Fee', '', 'Delta Group',
                         'Agency Fee'
                         ]
            else:
                stop_words = ['Delta', 'Direct', 'Group']
                words = campaign_name.split()
                result_words = [word for word in words if word not in stop_words]
                product_name = ' '.join(result_words)

                entry = ['Marketing', 'Sales Advertising', 'Traditional', '',
                         'Games/Prizes', '', 'Delta Group', product_name
                         ]
        elif re.search(r'Diamond(\s)*Lot', str(campaign_name), re.I):
            entry = ['Sales Ops', 'Vehicle Management', 'Professional Services',
                     'Photography', 'Photo Service', '', 'DiamondLot',
                     'DiamondLot Photo Overlay'
                     ]

    # Sponsor
    elif ttype == "Sponsor":
        entry = ['Marketing', set_func_area(dept), 'Traditional', '',
                 'Community Sponsorship', '', campaign_name, 'Sponsorship'
                 ]

    # Pre-roll
    elif re.search(r'(pre)(\s|-)*(rol+)', str(campaign_name), re.I):
        entry = ['Marketing', 'Sales Advertising', 'Online', 'Display', 'Video Pre-roll', '',
                 set_vendor(campaign_name, dmvendor, production_vendors), campaign_name
                 ]

    # OTT
    elif re.search(r'(^|\s)OTT(\s|$|/|:)', str(campaign_name), re.I):
        if re.search(r'Compulse', str(campaign_name), re.I):
            entry = ['Marketing', 'Sales Advertising', 'Online', 'Display',
                     'OTT', '', 'MAG', ':30 Compulse OTT'
                     ]
        elif re.search(r':15', str(campaign_name), re.I):
            entry = ['Marketing', 'Sales Advertising', 'Online', 'Display',
                     'OTT', '', 'MAG', 'OTT :15'
                     ]
        else:
            entry = ['Marketing', 'Sales Advertising', 'Online', 'Display',
                     'OTT', '', 'MAG', 'OTT :30'
                     ]

    # Digital/Traditional Radio
    elif re.search(r'iHeart', str(campaign_name), re.I):
        entry = ['Marketing', set_func_area(dept), 'Online',
                 'Mass Marketing', 'Digital Radio Advertising', '',
                 'iHeartMedia, Inc.', ':30 iHeart'
                 ]
    elif re.search(r'Pandora', str(campaign_name), re.I):
        entry = ['Marketing', set_func_area(dept), 'Online',
                 'Mass Marketing', 'Digital Radio Advertising',
                 'Pandora', 'Pandora', 'Pandora'
                 ]
    elif re.search(r'KXNO Hackfest', str(campaign_name), re.I):
        entry = ['Marketing', set_func_area(dept), 'Online',
                 'Mass Marketing', 'Digital Radio Advertising', '',
                 'MAG', 'KXNO Hackfest Digital Audio :60'
                 ]
    elif re.search(r'Spotify', str(campaign_name), re.I):
        entry = ['Marketing', set_func_area(dept), 'Online',
                 'Mass Marketing', 'Digital Radio Advertising',
                 'Spotify', 'MAG', 'Spotify'
                 ]
    elif ttype == "Radio" and not re.search(r'Spotify', str(campaign_name), re.I) and not \
            re.search(r'KXNO Hackfest', str(campaign_name), re.I) and not \
            re.search(r'Pandora', str(campaign_name), re.I) and not re.search(r'iHeart', str(campaign_name), re.I):
        entry = ['Marketing', set_func_area(dept), 'Traditional', 'Mass Marketing', 'Radio Advertising', '', 'MAG',
                 campaign_name
                 ]

    # YouTube TrueView
    elif re.search(r'YouTube', str(campaign_name), re.I):
        entry = ['Marketing', 'Sales Advertising', 'Online', 'Display',
                 'Video Pre-roll', 'YouTube', 'YouTube', 'TrueView'
                 ]

    # Car Media Group
    elif re.search(r'Car Media Group', str(dmvendor), re.I):
        if re.search(r'CRM Training', str(campaign_name), re.I):
            entry = ['Sales Ops', 'Training', 'Professional Services', '', 'CRM Training', '', 'Car Media Group',
                     'CRM Training/Workflow Setup'
                     ]

    # Mark Monitor
    elif re.search(r'Mark Monitor', str(campaign_name), re.I):
        entry = ['Marketing', set_func_area(dept), 'Online', 'Website', 'Website Hosting', '', 'MarkMonitor',
                 'Domain Renewal'
                 ]

    # Dex Yellow Pages
    elif re.search(r'Dex Media', str(campaign_name), re.I):
        if re.search(r'Digital', str(campaign_name)):
            entry = ['Marketing', set_func_area(dept), 'Online', 'Mass Marketing', 'Business Listing', '', 'DexYP',
                     'Dex Media Digital'
                     ]
        else:
            entry = ['Marketing', set_func_area(dept), 'Traditional', 'Mass Marketing', 'Business Listing', '', 'DexYP',
                     'Dex Media'
                     ]

    # Production
    elif ttype == "Production":
        # TV
        if re.search(r'(^|\s)TV(\s|$|/)', str(campaign_name), re.I) or \
                re.search(r'(^|\s)Camera(\s|$|/)', str(campaign_name), re.I) or \
                re.search(r'(^|\s)Film(\s|$|/)', str(campaign_name), re.I) or \
                re.search(r'(^|\s)Shoot(\s|$|/)', str(campaign_name), re.I) or \
                re.search(r'(^|\s)Suddenlink(\s|$|/)', str(campaign_name), re.I):
            entry = ['Marketing', 'Sales Advertising', 'Traditional', 'Mass Marketing', 'Television Advertising', '',
                     set_vendor(campaign_name, dmvendor, production_vendors), campaign_name
                     ]
        # Radio
        elif re.search(r'(^|\s)Radio(\s|$|/)', str(campaign_name), re.I) or \
                re.search(r'(^|\s)Audio(\s|$|/)', str(campaign_name), re.I):
            entry = ['Marketing', 'Sales Advertising', 'Traditional', 'Mass Marketing', 'Radio Advertising', '',
                     set_vendor(campaign_name, dmvendor, production_vendors), campaign_name
                     ]
        # Facebook
        elif re.search(r'(^|\s)Facebook(\s|$|/)', str(campaign_name), re.I):
            entry = ['Marketing', 'Sales Advertising', 'Online', 'Social - Paid', 'Social Ads', 'Facebook',
                     set_vendor(campaign_name, dmvendor, production_vendors), campaign_name
                     ]
        # Youtube/Pre-roll
        elif re.search(r'(^|\s)You(\s|-)*Tube(\s|$|/)', str(campaign_name), re.I) or \
                re.search(r'(pre)(\s|-)*(rol+)', str(campaign_name), re.I):
            entry = ['Marketing', 'Sales Advertising', 'Online', 'Display', 'Video Pre-roll', 'YouTube',
                     set_vendor(campaign_name, dmvendor, production_vendors), campaign_name
                     ]
        # Non-specific
        else:
            entry = ['Marketing', 'Sales Advertising', 'Professional Services', '', 'Content Production',
                     set_details(campaign_name), set_vendor(campaign_name, dmvendor, production_vendors), campaign_name
                     ]

    else:
        pass

    return entry


def update_translation_db():
    counter = 0
    passed_counter = 0

    confirmed_translations = {}
    for i in range(2, read_translation.max_row + 1):
        if read_translation['F{}'.format(i)].value:
            try:
                confirmed_translations[read_translation['O{}'.format(i)].value] = read_translation[
                    'P{}'.format(i)].value.split(";")
            except AttributeError:
                pass
        else:
            pass

    for i in range(2, translation.max_row+1):
        if translation['F{}'.format(i)].value is None:
            entry = auto_update(translation['A{}'.format(i)].value,
                                translation['B{}'.format(i)].value,
                                translation['C{}'.format(i)].value,
                                translation['D{}'.format(i)].value,
                                )
            if entry:
                translation['F{}'.format(i)].value = entry[0]
                translation['G{}'.format(i)].value = entry[1]
                translation['H{}'.format(i)].value = entry[2]
                translation['I{}'.format(i)].value = entry[3]
                translation['J{}'.format(i)].value = entry[4]
                translation['K{}'.format(i)].value = entry[5]
                translation['L{}'.format(i)].value = entry[6]
                translation['M{}'.format(i)].value = entry[7]
                translation['N{}'.format(i)].value = 1
                confirmed_translations[read_translation['O{}'.format(i)].value] = [data for data in entry]
                counter += 1
                print("Updated {} entries.".format(counter))

            else:
                key_list = []
                key_list.extend([
                    str(read_translation['A{}'.format(i)].value),
                    str(read_translation['B{}'.format(i)].value),
                    str(read_translation['C{}'.format(i)].value),
                    str(read_translation['D{}'.format(i)].value),
                    str(read_translation['E{}'.format(i)].value)
                ])
                key = ";".join(key_list)

                confirmed_key = get_close_matches(key, confirmed_translations.keys(), 1, cutoff=0.67)

                if len(confirmed_key) > 0:
                    dict_key = confirmed_key[0]
                    translation['F{}'.format(i)].value = confirmed_translations[dict_key][0]
                    translation['G{}'.format(i)].value = confirmed_translations[dict_key][1]
                    translation['H{}'.format(i)].value = confirmed_translations[dict_key][2]
                    translation['I{}'.format(i)].value = confirmed_translations[dict_key][3]
                    translation['J{}'.format(i)].value = confirmed_translations[dict_key][4]
                    translation['K{}'.format(i)].value = confirmed_translations[dict_key][5]
                    translation['L{}'.format(i)].value = confirmed_translations[dict_key][6]
                    translation['M{}'.format(i)].value = confirmed_translations[dict_key][7]
                    translation['N{}'.format(i)].value = 2
                    counter += 1
                    print("Updated {} entries.".format(counter))
                else:
                    passed_counter += 1
                    pass

                # if len(confirmed_key) > 0:
                #     dict_key = confirmed_key[0]
                #     translation['F{}'.format(i)].value = confirmed_translations[dict_key][0]
                #     translation['G{}'.format(i)].value = confirmed_translations[dict_key][1]
                #     translation['H{}'.format(i)].value = confirmed_translations[dict_key][2]
                #     translation['I{}'.format(i)].value = confirmed_translations[dict_key][3]
                #     translation['J{}'.format(i)].value = confirmed_translations[dict_key][4]
                #     translation['K{}'.format(i)].value = confirmed_translations[dict_key][5]
                #     translation['L{}'.format(i)].value = confirmed_translations[dict_key][6]
                #     translation['M{}'.format(i)].value = confirmed_translations[dict_key][7]
                #     translation['N{}'.format(i)].value = 2
                #     counter += 1
                #     print("Updated {} entries.".format(counter))
                # else:
                #     passed_counter += 1
                #     pass

    if passed_counter > 0:
        print("Failed to update {} entries.".format(passed_counter))
    elif passed_counter == 0:
        print("\nTranslation Database updated successfully! \n\nPlease wait for the message 'Process finished with "
              "exit code 0' to display before opening the Translation Database.")
    else:
        pass

    wb.save("{}.xlsx".format(translation_wb_name))


if __name__ == "__main__":
    update_translation_db()
